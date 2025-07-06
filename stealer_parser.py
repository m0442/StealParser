#!/usr/bin/env python3
"""
Enhanced Info Stealers Data Parser GUI
A modern, professional GUI with multiple export formats and advanced features.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import json
import re
import csv
import base64
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging
import queue
import webbrowser
import subprocess
import platform

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class InfoStealerParser:
    """Unified parser for various info stealer formats"""
    
    def __init__(self, base_path: str):
        self.base_path = Path(base_path)
        self.unified_data = {
            "metadata": {
                "parser_version": "2.0.0",
                "parsed_at": datetime.now().isoformat(),
                "total_sessions": 0,
                "stealer_types": []
            },
            "sessions": []
        }
        
    def parse_all(self) -> Dict[str, Any]:
        """Parse all stealer data in the base directory"""
        logger.info(f"Starting to parse stealer data from: {self.base_path}")
        
        # Get all stealer directories
        stealer_dirs = [d for d in self.base_path.iterdir() if d.is_dir() and not d.name.startswith('.')]
        
        for stealer_dir in stealer_dirs:
            stealer_type = stealer_dir.name
            logger.info(f"Processing stealer type: {stealer_type}")
            
            # Parse each stealer type
            if stealer_type == "Redline":
                self._parse_redline(stealer_dir)
            elif stealer_type == "Mystic Stealer":
                self._parse_mystic(stealer_dir)
            elif stealer_type == "Meta Stealer":
                self._parse_meta(stealer_dir)
            elif stealer_type == "Luma Stealer":
                self._parse_luma(stealer_dir)
            elif stealer_type == "LumaC2 Stealer":
                self._parse_lumac2(stealer_dir)
            elif stealer_type == "Raccoon":
                self._parse_raccoon(stealer_dir)
            elif stealer_type == "Old Redline":
                self._parse_old_redline(stealer_dir)
            elif stealer_type == "Stealc Stealer":
                self._parse_stealc(stealer_dir)
            elif stealer_type == "Vider":
                self._parse_vider(stealer_dir)
            elif stealer_type == "Ununkown Stealer":
                self._parse_unknown(stealer_dir)
            elif stealer_type == "Dark Crystal RAT Stealer":
                self._parse_dark_crystal(stealer_dir)
            else:
                logger.warning(f"Unknown stealer type: {stealer_type}")
                
        self.unified_data["metadata"]["total_sessions"] = len(self.unified_data["sessions"])
        return self.unified_data
    
    def _parse_redline(self, stealer_dir: Path):
        """Parse Redline stealer data"""
        for session_dir in stealer_dir.iterdir():
            if not session_dir.is_dir():
                continue
                
            session_data = {
                "stealer_type": "Redline",
                "session_id": session_dir.name,
                "system_info": {},
                "passwords": [],
                "cookies": [],
                "autofills": [],
                "files": [],
                "screenshots": [],
                "errors": []
            }
            
            try:
                # Parse system information
                user_info_file = session_dir / "UserInformation.txt"
                if user_info_file.exists():
                    session_data["system_info"] = self._parse_redline_system_info(user_info_file)
                
                # Parse passwords
                passwords_file = session_dir / "Passwords.txt"
                if passwords_file.exists():
                    session_data["passwords"] = self._parse_redline_passwords(passwords_file)
                
                # Parse cookies
                cookies_dir = session_dir / "Cookies"
                if cookies_dir.exists():
                    session_data["cookies"] = self._parse_cookies(cookies_dir)
                
                # Parse autofills
                autofills_dir = session_dir / "Autofills"
                if autofills_dir.exists():
                    session_data["autofills"] = self._parse_autofills(autofills_dir)
                
                # Parse other files
                for file_path in session_dir.iterdir():
                    if file_path.is_file():
                        if file_path.suffix.lower() in ['.jpg', '.png', '.jpeg']:
                            session_data["screenshots"].append({
                                "filename": file_path.name,
                                "size": file_path.stat().st_size,
                                "path": str(file_path.relative_to(self.base_path))
                            })
                        elif file_path.suffix.lower() == '.txt':
                            session_data["files"].append({
                                "filename": file_path.name,
                                "size": file_path.stat().st_size,
                                "path": str(file_path.relative_to(self.base_path))
                            })
                
                self.unified_data["sessions"].append(session_data)
                
            except Exception as e:
                logger.error(f"Error parsing Redline session {session_dir.name}: {e}")
                session_data["errors"].append(str(e))
                self.unified_data["sessions"].append(session_data)
    
    def _parse_redline_system_info(self, file_path: Path) -> Dict[str, Any]:
        """Parse Redline system information"""
        system_info = {}
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                
            # Extract key information using regex patterns
            patterns = {
                'ip': r'IP:\s*([^\n]+)',
                'username': r'UserName:\s*([^\n]+)',
                'country': r'Country:\s*([^\n]+)',
                'location': r'Location:\s*([^\n]+)',
                'zip_code': r'Zip Code:\s*([^\n]+)',
                'hwid': r'HWID:\s*([^\n]+)',
                'language': r'Current Language:\s*([^\n]+)',
                'screen_size': r'ScreenSize:\s*([^\n]+)',
                'timezone': r'TimeZone:\s*([^\n]+)',
                'os': r'Operation System:\s*([^\n]+)',
                'log_date': r'Log date:\s*([^\n]+)'
            }
            
            for key, pattern in patterns.items():
                match = re.search(pattern, content)
                if match:
                    system_info[key] = match.group(1).strip()
            
            # Extract hardware information
            hw_section = re.search(r'Hardwares:\s*(.*?)(?=\n\n|\nAnti-Viruses:|$)', content, re.DOTALL)
            if hw_section:
                system_info['hardware'] = hw_section.group(1).strip().split('\n')
            
            # Extract antivirus information
            av_section = re.search(r'Anti-Viruses:\s*(.*?)(?=\n\n|$)', content, re.DOTALL)
            if av_section:
                system_info['antivirus'] = av_section.group(1).strip().split('\n')
                
        except Exception as e:
            logger.error(f"Error parsing system info from {file_path}: {e}")
            
        return system_info
    
    def _parse_redline_passwords(self, file_path: Path) -> List[Dict[str, str]]:
        """Parse Redline passwords"""
        passwords = []
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Split by password entries
            entries = content.split('===============')
            
            for entry in entries:
                if not entry.strip():
                    continue
                    
                password_data = {}
                
                # Extract URL, username, password, and application
                url_match = re.search(r'URL:\s*(.+)', entry)
                if url_match:
                    password_data['url'] = url_match.group(1).strip()
                
                username_match = re.search(r'Username:\s*(.+)', entry)
                if username_match:
                    password_data['username'] = username_match.group(1).strip()
                
                password_match = re.search(r'Password:\s*(.+)', entry)
                if password_match:
                    password_data['password'] = password_match.group(1).strip()
                
                app_match = re.search(r'Application:\s*(.+)', entry)
                if app_match:
                    password_data['application'] = app_match.group(1).strip()
                
                if password_data:
                    passwords.append(password_data)
                    
        except Exception as e:
            logger.error(f"Error parsing passwords from {file_path}: {e}")
            
        return passwords
    
    def _parse_mystic(self, stealer_dir: Path):
        """Parse Mystic Stealer data"""
        for session_dir in stealer_dir.iterdir():
            if not session_dir.is_dir():
                continue
                
            session_data = {
                "stealer_type": "Mystic Stealer",
                "session_id": session_dir.name,
                "system_info": {},
                "passwords": [],
                "cookies": [],
                "autofills": [],
                "credit_cards": [],
                "telegram": [],
                "wallets": [],
                "errors": []
            }
            
            try:
                # Parse system information
                sys_info_file = session_dir / "SystemInformation.txt"
                if sys_info_file.exists():
                    session_data["system_info"] = self._parse_mystic_system_info(sys_info_file)
                
                # Parse passwords
                passwords_file = session_dir / "Passwords.txt"
                if passwords_file.exists():
                    session_data["passwords"] = self._parse_mystic_passwords(passwords_file)
                
                # Parse cookies
                cookies_dir = session_dir / "Cookies"
                if cookies_dir.exists():
                    session_data["cookies"] = self._parse_cookies(cookies_dir)
                
                # Parse autofills
                autofills_dir = session_dir / "Autofills"
                if autofills_dir.exists():
                    session_data["autofills"] = self._parse_autofills(autofills_dir)
                
                # Parse credit cards
                credit_cards_dir = session_dir / "CreditCards"
                if credit_cards_dir.exists():
                    session_data["credit_cards"] = self._parse_credit_cards(credit_cards_dir)
                
                # Parse Telegram data
                telegram_dir = session_dir / "Telegram"
                if telegram_dir.exists():
                    session_data["telegram"] = self._parse_telegram(telegram_dir)
                
                # Parse wallets
                wallets_dir = session_dir / "Wallets"
                if wallets_dir.exists():
                    session_data["wallets"] = self._parse_wallets(wallets_dir)
                
                self.unified_data["sessions"].append(session_data)
                
            except Exception as e:
                logger.error(f"Error parsing Mystic session {session_dir.name}: {e}")
                session_data["errors"].append(str(e))
                self.unified_data["sessions"].append(session_data)
    
    def _parse_mystic_system_info(self, file_path: Path) -> Dict[str, Any]:
        """Parse Mystic Stealer system information"""
        system_info = {}
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                
            patterns = {
                'ip': r'IP:\s*([^\n]+)',
                'username': r'UserName:\s*([^\n]+)',
                'computer_name': r'ComputerName:\s*([^\n]+)',
                'country': r'Country:\s*([^\n]+)',
                'location': r'Location:\s*([^\n]+)',
                'zip_code': r'Zip code:\s*([^\n]+)',
                'timezone': r'TimeZone:\s*([^\n]+)',
                'hwid': r'HWID:\s*([^\n]+)',
                'language': r'Current language:\s*([^\n]+)',
                'screen_size': r'ScreenSize:\s*([^\n]+)',
                'os': r'Operation System:\s*([^\n]+)'
            }
            
            for key, pattern in patterns.items():
                match = re.search(pattern, content)
                if match:
                    system_info[key] = match.group(1).strip()
            
            # Extract hardware information
            hw_section = re.search(r'Hardwares:\s*(.*?)(?=\n\n|$)', content, re.DOTALL)
            if hw_section:
                system_info['hardware'] = hw_section.group(1).strip().split('\n')
                
        except Exception as e:
            logger.error(f"Error parsing Mystic system info from {file_path}: {e}")
            
        return system_info
    
    def _parse_mystic_passwords(self, file_path: Path) -> List[Dict[str, str]]:
        """Parse Mystic Stealer passwords"""
        passwords = []
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Split by password entries
            entries = content.split('===============')
            
            for entry in entries:
                if not entry.strip():
                    continue
                    
                password_data = {}
                
                # Extract URL, username, password, and application
                url_match = re.search(r'URL:\s*(.+)', entry)
                if url_match:
                    password_data['url'] = url_match.group(1).strip()
                
                username_match = re.search(r'Username:\s*(.+)', entry)
                if username_match:
                    password_data['username'] = username_match.group(1).strip()
                
                password_match = re.search(r'Password:\s*(.+)', entry)
                if password_match:
                    password_data['password'] = password_match.group(1).strip()
                
                app_match = re.search(r'Application:\s*(.+)', entry)
                if app_match:
                    password_data['application'] = app_match.group(1).strip()
                
                if password_data:
                    passwords.append(password_data)
                    
        except Exception as e:
            logger.error(f"Error parsing Mystic passwords from {file_path}: {e}")
            
        return passwords
    
    def _parse_luma(self, stealer_dir: Path):
        """Parse Luma Stealer data"""
        for session_dir in stealer_dir.iterdir():
            if not session_dir.is_dir():
                continue
                
            session_data = {
                "stealer_type": "Luma Stealer",
                "session_id": session_dir.name,
                "system_info": {},
                "passwords": [],
                "cookies": [],
                "applications": [],
                "errors": []
            }
            
            try:
                # Parse system information
                system_file = session_dir / "System.txt"
                if system_file.exists():
                    session_data["system_info"] = self._parse_luma_system_info(system_file)
                
                # Parse passwords
                passwords_file = session_dir / "All Passwords.txt"
                if passwords_file.exists():
                    session_data["passwords"] = self._parse_luma_passwords(passwords_file)
                
                # Parse cookies
                cookies_dir = session_dir / "Cookies"
                if cookies_dir.exists():
                    session_data["cookies"] = self._parse_cookies(cookies_dir)
                
                # Parse applications
                apps_dir = session_dir / "Applications"
                if apps_dir.exists():
                    session_data["applications"] = self._parse_applications(apps_dir)
                
                self.unified_data["sessions"].append(session_data)
                
            except Exception as e:
                logger.error(f"Error parsing Luma session {session_dir.name}: {e}")
                session_data["errors"].append(str(e))
                self.unified_data["sessions"].append(session_data)
    
    def _parse_luma_system_info(self, file_path: Path) -> Dict[str, Any]:
        """Parse Luma Stealer system information"""
        system_info = {}
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                
            patterns = {
                'pc': r'- PC:\s*([^\n]+)',
                'user': r'- User:\s*([^\n]+)',
                'domain': r'- Domain:\s*([^\n]+)',
                'workgroup': r'- Workgroup:\s*([^\n]+)',
                'computer_name': r'- ComputerNameDnsHostname:\s*([^\n]+)',
                'os_version': r'- OS Version:\s*([^\n]+)',
                'hwid': r'- HWID:\s*([^\n]+)',
                'screen_resolution': r'- Screen Resoluton:\s*([^\n]+)',
                'language': r'- Language:\s*([^\n]+)',
                'cpu': r'- CPU Name:\s*([^\n]+)',
                'gpu': r'- GPU:\s*([^\n]+)',
                'ram': r'- Physical Installed Memory:\s*([^\n]+)',
                'ip': r'- IP Address:\s*([^\n]+)',
                'country': r'- Country:\s*([^\n]+)'
            }
            
            for key, pattern in patterns.items():
                match = re.search(pattern, content)
                if match:
                    system_info[key] = match.group(1).strip()
                    
        except Exception as e:
            logger.error(f"Error parsing Luma system info from {file_path}: {e}")
            
        return system_info
    
    def _parse_luma_passwords(self, file_path: Path) -> List[Dict[str, str]]:
        """Parse Luma Stealer passwords"""
        passwords = []
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Split by password entries
            entries = content.split('\n\n')
            
            for entry in entries:
                if not entry.strip():
                    continue
                    
                password_data = {}
                
                # Extract software, URL, username, and password
                software_match = re.search(r'SOFT:\s*(.+)', entry)
                if software_match:
                    password_data['software'] = software_match.group(1).strip()
                
                url_match = re.search(r'URL:\s*(.+)', entry)
                if url_match:
                    password_data['url'] = url_match.group(1).strip()
                
                user_match = re.search(r'USER:\s*(.+)', entry)
                if user_match:
                    password_data['username'] = user_match.group(1).strip()
                
                pass_match = re.search(r'PASS:\s*(.+)', entry)
                if pass_match:
                    password_data['password'] = pass_match.group(1).strip()
                
                if password_data:
                    passwords.append(password_data)
                    
        except Exception as e:
            logger.error(f"Error parsing Luma passwords from {file_path}: {e}")
            
        return passwords
    
    def _parse_raccoon(self, stealer_dir: Path):
        """Parse Raccoon stealer data"""
        for session_dir in stealer_dir.iterdir():
            if not session_dir.is_dir():
                continue
                
            session_data = {
                "stealer_type": "Raccoon",
                "session_id": session_dir.name,
                "system_info": {},
                "passwords": [],
                "browsers": [],
                "errors": []
            }
            
            try:
                # Parse system information
                sys_info_file = session_dir / "System Info.txt"
                if sys_info_file.exists():
                    session_data["system_info"] = self._parse_raccoon_system_info(sys_info_file)
                
                # Parse passwords
                passwords_file = session_dir / "passwords.txt"
                if passwords_file.exists():
                    session_data["passwords"] = self._parse_raccoon_passwords(passwords_file)
                
                # Parse browsers
                browsers_dir = session_dir / "browsers"
                if browsers_dir.exists():
                    session_data["browsers"] = self._parse_browsers(browsers_dir)
                
                self.unified_data["sessions"].append(session_data)
                
            except Exception as e:
                logger.error(f"Error parsing Raccoon session {session_dir.name}: {e}")
                session_data["errors"].append(str(e))
                self.unified_data["sessions"].append(session_data)
    
    def _parse_raccoon_system_info(self, file_path: Path) -> Dict[str, Any]:
        """Parse Raccoon system information"""
        system_info = {}
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                
            patterns = {
                'system_language': r'System Language:\s*([^\n]+)',
                'system_timezone': r'System TimeZone:\s*([^\n]+)',
                'ip': r'IP:\s*([^\n]+)',
                'location': r'Location:\s*([^\n]+)',
                'computer_name': r'ComputerName:\s*([^\n]+)',
                'username': r'Username:\s*([^\n]+)',
                'windows_version': r'Windows version:\s*([^\n]+)',
                'product_name': r'Product name:\s*([^\n]+)',
                'system_arch': r'System arch:\s*([^\n]+)',
                'cpu': r'CPU:\s*([^\n]+)',
                'ram': r'RAM:\s*([^\n]+)',
                'screen_resolution': r'Screen resolution:\s*([^\n]+)'
            }
            
            for key, pattern in patterns.items():
                match = re.search(pattern, content)
                if match:
                    system_info[key] = match.group(1).strip()
            
            # Extract display devices
            display_section = re.search(r'Display devices:\s*(.*?)(?=\n\n|\nInstalled Apps:|$)', content, re.DOTALL)
            if display_section:
                system_info['display_devices'] = display_section.group(1).strip().split('\n')
            
            # Extract installed apps
            apps_section = re.search(r'Installed Apps:\s*(.*?)(?=\n\n|$)', content, re.DOTALL)
            if apps_section:
                system_info['installed_apps'] = apps_section.group(1).strip().split('\n')
                
        except Exception as e:
            logger.error(f"Error parsing Raccoon system info from {file_path}: {e}")
            
        return system_info
    
    def _parse_raccoon_passwords(self, file_path: Path) -> List[Dict[str, str]]:
        """Parse Raccoon passwords"""
        passwords = []
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Split by lines and parse each entry
            lines = content.strip().split('\n')
            for line in lines:
                if ':' in line:
                    parts = line.split(':', 1)
                    if len(parts) == 2:
                        passwords.append({
                            'url': parts[0].strip(),
                            'password': parts[1].strip()
                        })
                        
        except Exception as e:
            logger.error(f"Error parsing Raccoon passwords from {file_path}: {e}")
            
        return passwords
    
    def _parse_cookies(self, cookies_dir: Path) -> List[Dict[str, Any]]:
        """Parse cookies from various formats"""
        cookies = []
        
        try:
            for cookie_file in cookies_dir.iterdir():
                if cookie_file.is_file() and cookie_file.suffix.lower() == '.txt':
                    cookie_data = {
                        'filename': cookie_file.name,
                        'size': cookie_file.stat().st_size,
                        'entries': []
                    }
                    
                    with open(cookie_file, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                    
                    # Parse Netscape format cookies
                    lines = content.strip().split('\n')
                    for line in lines:
                        if line and not line.startswith('#'):
                            parts = line.split('\t')
                            if len(parts) >= 7:
                                cookie_data['entries'].append({
                                    'domain': parts[0],
                                    'path': parts[2],
                                    'secure': parts[3] == 'TRUE',
                                    'expiry': parts[4],
                                    'name': parts[5],
                                    'value': parts[6]
                                })
                    
                    cookies.append(cookie_data)
                    
        except Exception as e:
            logger.error(f"Error parsing cookies from {cookies_dir}: {e}")
            
        return cookies
    
    def _parse_autofills(self, autofills_dir: Path) -> List[Dict[str, Any]]:
        """Parse autofill data"""
        autofills = []
        
        try:
            for autofill_file in autofills_dir.iterdir():
                if autofill_file.is_file() and autofill_file.suffix.lower() == '.txt':
                    autofill_data = {
                        'filename': autofill_file.name,
                        'size': autofill_file.stat().st_size,
                        'entries': []
                    }
                    
                    with open(autofill_file, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                    
                    # Parse autofill entries
                    lines = content.strip().split('\n')
                    for line in lines:
                        if line and ':' in line:
                            parts = line.split(':', 1)
                            if len(parts) == 2:
                                autofill_data['entries'].append({
                                    'field': parts[0].strip(),
                                    'value': parts[1].strip()
                                })
                    
                    autofills.append(autofill_data)
                    
        except Exception as e:
            logger.error(f"Error parsing autofills from {autofills_dir}: {e}")
            
        return autofills
    
    def _parse_credit_cards(self, credit_cards_dir: Path) -> List[Dict[str, Any]]:
        """Parse credit card data"""
        credit_cards = []
        
        try:
            for cc_file in credit_cards_dir.iterdir():
                if cc_file.is_file() and cc_file.suffix.lower() == '.txt':
                    cc_data = {
                        'filename': cc_file.name,
                        'size': cc_file.stat().st_size,
                        'entries': []
                    }
                    
                    with open(cc_file, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                    
                    # Parse credit card entries
                    lines = content.strip().split('\n')
                    for line in lines:
                        if line and ':' in line:
                            parts = line.split(':', 1)
                            if len(parts) == 2:
                                cc_data['entries'].append({
                                    'field': parts[0].strip(),
                                    'value': parts[1].strip()
                                })
                    
                    credit_cards.append(cc_data)
                    
        except Exception as e:
            logger.error(f"Error parsing credit cards from {credit_cards_dir}: {e}")
            
        return credit_cards
    
    def _parse_telegram(self, telegram_dir: Path) -> List[Dict[str, Any]]:
        """Parse Telegram data"""
        telegram_data = []
        
        try:
            for item in telegram_dir.iterdir():
                if item.is_file():
                    telegram_data.append({
                        'filename': item.name,
                        'size': item.stat().st_size,
                        'path': str(item.relative_to(self.base_path))
                    })
                elif item.is_dir():
                    telegram_data.append({
                        'directory': item.name,
                        'contents': [f.name for f in item.iterdir()]
                    })
                    
        except Exception as e:
            logger.error(f"Error parsing Telegram data from {telegram_dir}: {e}")
            
        return telegram_data
    
    def _parse_wallets(self, wallets_dir: Path) -> List[Dict[str, Any]]:
        """Parse wallet data"""
        wallets = []
        
        try:
            for wallet_dir in wallets_dir.iterdir():
                if wallet_dir.is_dir():
                    wallet_data = {
                        'wallet_type': wallet_dir.name,
                        'files': []
                    }
                    
                    for file_path in wallet_dir.iterdir():
                        if file_path.is_file():
                            wallet_data['files'].append({
                                'filename': file_path.name,
                                'size': file_path.stat().st_size,
                                'path': str(file_path.relative_to(self.base_path))
                            })
                    
                    wallets.append(wallet_data)
                    
        except Exception as e:
            logger.error(f"Error parsing wallets from {wallets_dir}: {e}")
            
        return wallets
    
    def _parse_applications(self, apps_dir: Path) -> List[Dict[str, Any]]:
        """Parse application data"""
        applications = []
        
        try:
            for app_dir in apps_dir.iterdir():
                if app_dir.is_dir():
                    app_data = {
                        'application': app_dir.name,
                        'files': []
                    }
                    
                    for file_path in app_dir.rglob('*'):
                        if file_path.is_file():
                            app_data['files'].append({
                                'filename': file_path.name,
                                'size': file_path.stat().st_size,
                                'path': str(file_path.relative_to(self.base_path))
                            })
                    
                    applications.append(app_data)
                    
        except Exception as e:
            logger.error(f"Error parsing applications from {apps_dir}: {e}")
            
        return applications
    
    def _parse_browsers(self, browsers_dir: Path) -> List[Dict[str, Any]]:
        """Parse browser data"""
        browsers = []
        
        try:
            for browser_dir in browsers_dir.iterdir():
                if browser_dir.is_dir():
                    browser_data = {
                        'browser': browser_dir.name,
                        'files': []
                    }
                    
                    for file_path in browser_dir.rglob('*'):
                        if file_path.is_file():
                            browser_data['files'].append({
                                'filename': file_path.name,
                                'size': file_path.stat().st_size,
                                'path': str(file_path.relative_to(self.base_path))
                            })
                    
                    browsers.append(browser_data)
                    
        except Exception as e:
            logger.error(f"Error parsing browsers from {browsers_dir}: {e}")
            
        return browsers
    
    def _parse_meta(self, stealer_dir: Path):
        """Parse Meta Stealer data"""
        # Implementation for Meta Stealer
        pass
    
    def _parse_lumac2(self, stealer_dir: Path):
        """Parse LumaC2 Stealer data"""
        # Implementation for LumaC2 Stealer
        pass
    
    def _parse_old_redline(self, stealer_dir: Path):
        """Parse Old Redline data"""
        # Implementation for Old Redline
        pass
    
    def _parse_stealc(self, stealer_dir: Path):
        """Parse Stealc Stealer data"""
        # Implementation for Stealc Stealer
        pass
    
    def _parse_vider(self, stealer_dir: Path):
        """Parse Vider stealer data"""
        # Implementation for Vider stealer
        pass
    
    def _parse_unknown(self, stealer_dir: Path):
        """Parse Unknown stealer data"""
        # Implementation for unknown stealers
        pass
    
    def _parse_dark_crystal(self, stealer_dir: Path):
        """Parse Dark Crystal RAT Stealer data"""
        # Implementation for Dark Crystal RAT Stealer
        pass

class DataExporter:
    """Handles multiple export formats"""
    
    @staticmethod
    def export_json(data: Dict[str, Any], output_path: str):
        """Export data to JSON format"""
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            return True, f"JSON exported successfully to {output_path}"
        except Exception as e:
            return False, f"Error exporting JSON: {str(e)}"
    
    @staticmethod
    def export_csv(data: Dict[str, Any], output_path: str):
        """Export passwords and system info to CSV format"""
        try:
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write header
                writer.writerow(['Stealer Type', 'Session ID', 'URL', 'Username', 'Password', 'Application', 'IP', 'Country', 'OS'])
                
                # Write data
                for session in data.get('sessions', []):
                    stealer_type = session.get('stealer_type', '')
                    session_id = session.get('session_id', '')
                    system_info = session.get('system_info', {})
                    ip = system_info.get('ip', '')
                    country = system_info.get('country', '')
                    os_info = system_info.get('os', '')
                    
                    for password in session.get('passwords', []):
                        writer.writerow([
                            stealer_type,
                            session_id,
                            password.get('url', ''),
                            password.get('username', ''),
                            password.get('password', ''),
                            password.get('application', ''),
                            ip,
                            country,
                            os_info
                        ])
            
            return True, f"CSV exported successfully to {output_path}"
        except Exception as e:
            return False, f"Error exporting CSV: {str(e)}"
    
    @staticmethod
    def export_excel(data: Dict[str, Any], output_path: str):
        """Export data to Excel format with multiple sheets"""
        try:
            # Try to import openpyxl, if not available, fall back to CSV
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                return False, "openpyxl not installed. Please install with: pip install openpyxl"
            
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Summary sheet
            summary_ws = wb.create_sheet("Summary")
            summary_ws['A1'] = "Info Stealers Data Parser - Summary Report"
            summary_ws['A1'].font = Font(bold=True, size=14)
            
            summary_ws['A3'] = "Generated:"
            summary_ws['B3'] = data.get('metadata', {}).get('parsed_at', '')
            summary_ws['A4'] = "Total Sessions:"
            summary_ws['B4'] = data.get('metadata', {}).get('total_sessions', 0)
            summary_ws['A5'] = "Parser Version:"
            summary_ws['B5'] = data.get('metadata', {}).get('parser_version', '')
            
            # Passwords sheet
            passwords_ws = wb.create_sheet("Passwords")
            headers = ['Stealer Type', 'Session ID', 'URL', 'Username', 'Password', 'Application', 'IP', 'Country', 'OS']
            for col, header in enumerate(headers, 1):
                cell = passwords_ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            row = 2
            for session in data.get('sessions', []):
                stealer_type = session.get('stealer_type', '')
                session_id = session.get('session_id', '')
                system_info = session.get('system_info', {})
                ip = system_info.get('ip', '')
                country = system_info.get('country', '')
                os_info = system_info.get('os', '')
                
                for password in session.get('passwords', []):
                    passwords_ws.cell(row=row, column=1, value=stealer_type)
                    passwords_ws.cell(row=row, column=2, value=session_id)
                    passwords_ws.cell(row=row, column=3, value=password.get('url', ''))
                    passwords_ws.cell(row=row, column=4, value=password.get('username', ''))
                    passwords_ws.cell(row=row, column=5, value=password.get('password', ''))
                    passwords_ws.cell(row=row, column=6, value=password.get('application', ''))
                    passwords_ws.cell(row=row, column=7, value=ip)
                    passwords_ws.cell(row=row, column=8, value=country)
                    passwords_ws.cell(row=row, column=9, value=os_info)
                    row += 1
            
            # System Info sheet
            sysinfo_ws = wb.create_sheet("System Info")
            sysinfo_headers = ['Stealer Type', 'Session ID', 'IP', 'Country', 'Location', 'Username', 'Computer Name', 'OS', 'HWID', 'Language']
            for col, header in enumerate(sysinfo_headers, 1):
                cell = sysinfo_ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            row = 2
            for session in data.get('sessions', []):
                stealer_type = session.get('stealer_type', '')
                session_id = session.get('session_id', '')
                system_info = session.get('system_info', {})
                
                sysinfo_ws.cell(row=row, column=1, value=stealer_type)
                sysinfo_ws.cell(row=row, column=2, value=session_id)
                sysinfo_ws.cell(row=row, column=3, value=system_info.get('ip', ''))
                sysinfo_ws.cell(row=row, column=4, value=system_info.get('country', ''))
                sysinfo_ws.cell(row=row, column=5, value=system_info.get('location', ''))
                sysinfo_ws.cell(row=row, column=6, value=system_info.get('username', ''))
                sysinfo_ws.cell(row=row, column=7, value=system_info.get('computer_name', ''))
                sysinfo_ws.cell(row=row, column=8, value=system_info.get('os', ''))
                sysinfo_ws.cell(row=row, column=9, value=system_info.get('hwid', ''))
                sysinfo_ws.cell(row=row, column=10, value=system_info.get('language', ''))
                row += 1
            
            # Statistics sheet
            stats_ws = wb.create_sheet("Statistics")
            stats_ws['A1'] = "Parsing Statistics"
            stats_ws['A1'].font = Font(bold=True, size=14)
            
            stats_ws['A3'] = "Stealer Types Found:"
            stealer_types = set()
            total_passwords = 0
            total_cookies = 0
            
            for session in data.get('sessions', []):
                stealer_types.add(session.get('stealer_type', ''))
                total_passwords += len(session.get('passwords', []))
                total_cookies += len(session.get('cookies', []))
            
            stats_ws['B3'] = ', '.join(stealer_types)
            stats_ws['A4'] = "Total Passwords:"
            stats_ws['B4'] = total_passwords
            stats_ws['A5'] = "Total Cookie Files:"
            stats_ws['B5'] = total_cookies
            
            wb.save(output_path)
            return True, f"Excel file exported successfully to {output_path}"
            
        except Exception as e:
            return False, f"Error exporting Excel: {str(e)}"
    
    @staticmethod
    def export_html(data: Dict[str, Any], output_path: str):
        """Export data to HTML format with styling"""
        try:
            html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Info Stealers Data Parser Report</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #1e1e1e;
            color: #ffffff;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: #2b2b2b;
            border-radius: 10px;
            padding: 20px;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.3);
        }}
        .header {{
            text-align: center;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 2px solid #4CAF50;
        }}
        .header h1 {{
            color: #4CAF50;
            margin: 0;
            font-size: 2.5em;
        }}
        .summary {{
            background-color: #3c3c3c;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
        .summary h2 {{
            color: #4CAF50;
            margin-top: 0;
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-top: 15px;
        }}
        .stat-card {{
            background-color: #4a4a4a;
            padding: 15px;
            border-radius: 5px;
            text-align: center;
        }}
        .stat-number {{
            font-size: 2em;
            font-weight: bold;
            color: #4CAF50;
        }}
        .stat-label {{
            color: #cccccc;
            margin-top: 5px;
        }}
        .section {{
            margin-bottom: 30px;
        }}
        .section h2 {{
            color: #4CAF50;
            border-bottom: 1px solid #4CAF50;
            padding-bottom: 10px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
            background-color: #3c3c3c;
            border-radius: 5px;
            overflow: hidden;
        }}
        th {{
            background-color: #4CAF50;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            padding: 10px 12px;
            border-bottom: 1px solid #4a4a4a;
        }}
        tr:hover {{
            background-color: #4a4a4a;
        }}
        .password-row {{
            background-color: #2d2d2d;
        }}
        .footer {{
            text-align: center;
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #4a4a4a;
            color: #888888;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1> Info Stealers Data Parser</h1>
            <p>Comprehensive Analysis Report</p>
        </div>
        
        <div class="summary">
            <h2>📊 Executive Summary</h2>
            <div class="stats-grid">
                <div class="stat-card">
                    <div class="stat-number">{data.get('metadata', {}).get('total_sessions', 0)}</div>
                    <div class="stat-label">Total Sessions</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{len(set(s.get('stealer_type', '') for s in data.get('sessions', [])))}</div>
                    <div class="stat-label">Stealer Types</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{sum(len(s.get('passwords', [])) for s in data.get('sessions', []))}</div>
                    <div class="stat-label">Total Passwords</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{len(set(s.get('system_info', {}).get('country', '') for s in data.get('sessions', []) if s.get('system_info', {}).get('country', '')))}</div>
                    <div class="stat-label">Total Countries</div>
                </div>
            </div>
        </div>
        
        <div class="section">
            <h2>📋 Detailed Session Information</h2>
            <table>
                <thead>
                    <tr>
                        <th>Stealer Type</th>
                        <th>Session ID</th>
                        <th>System Info</th>
                        <th>Passwords</th>
                        <th>Cookies</th>
                        <th>Autofills</th>
                        <th>Screenshots</th>
                        <th>Files</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join([f"<tr><td>{session.get('stealer_type', '')}</td><td>{session.get('session_id', '')}</td><td>{json.dumps(session.get('system_info', {}))}</td><td>{len(session.get('passwords', []))}</td><td>{len(session.get('cookies', []))}</td><td>{len(session.get('autofills', []))}</td><td>{len(session.get('screenshots', []))}</td><td>{len(session.get('files', []))}</td></tr>" for session in data.get('sessions', [])])}
                </tbody>
            </table>
        </div>
        
        <div class="footer">
            <p>Generated at: {data.get('metadata', {}).get('parsed_at', '')}</p>
        </div>
    </div>
</body>
</html>
"""
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            return True, f"HTML exported successfully to {output_path}"
        except Exception as e:
            return False, f"Error exporting HTML: {str(e)}"